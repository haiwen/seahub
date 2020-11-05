# Copyright (c) 2012-2016 Seafile Ltd.
import logging
import openpyxl

logger = logging.getLogger(__name__)

def write_xls(sheet_name, head, data_list):
    """write listed data into excel
    """

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
    except Exception as e:
        logger.error(e)
        return None

    ws.title = sheet_name

    row_num = 0

    # write table head
    for col_num in range(len(head)):
        c = ws.cell(row = row_num + 1, column = col_num + 1)
        c.value = head[col_num]

    # write table data
    for row in data_list:
        row_num += 1
        for col_num in range(len(row)):
            c = ws.cell(row = row_num + 1, column = col_num + 1)
            c.value = row[col_num]

    return wb
